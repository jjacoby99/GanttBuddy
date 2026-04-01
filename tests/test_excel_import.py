from __future__ import annotations

import datetime as dt
import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from logic.load_project import DataColumn, ExcelParameters, ExcelProjectLoader
from models.project_type import ProjectType


def _default_excel_parameters(start_row: int = 8) -> ExcelParameters:
    return ExcelParameters(
        start_row=start_row,
        columns=[
            DataColumn(name="ACTIVITY", column=2),
            DataColumn(name="PLANNED DURATION (HOURS)", column=3),
            DataColumn(name="PLANNED START", column=4),
            DataColumn(name="PLANNED END", column=5),
            DataColumn(name="ACTUAL DURATION", column=7),
            DataColumn(name="ACTUAL START", column=8),
            DataColumn(name="ACTUAL END", column=9),
            DataColumn(name="NOTES", column=10),
            DataColumn(name="PREDECESSOR", column=11),
            DataColumn(name="UUID", column=12),
            DataColumn(name="PLANNED", column=13),
        ],
    )


def _build_excel_fixture() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Schedule"
    ws["A5"] = "Fixture Project"

    headers = [
        "IGNORE",
        "ACTIVITY",
        "PLANNED DURATION (HOURS)",
        "PLANNED START",
        "PLANNED END",
        "IGNORE 2",
        "ACTUAL DURATION",
        "ACTUAL START",
        "ACTUAL END",
        "NOTES",
        "PREDECESSOR",
        "UUID",
        "PLANNED",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=7, column=idx, value=header)

    # Row 8 is dropped by the loader, matching the template layout.
    ws.cell(row=8, column=2, value="Template spacer")

    schedule_rows = {
        9: ["Phase 1", "1 Day", dt.datetime(2026, 2, 17, 7), dt.datetime(2026, 2, 18, 7), None, None, None, "", "", "", True],
        10: ["Task A", 2, dt.datetime(2026, 2, 17, 7), dt.datetime(2026, 2, 17, 9), None, None, None, "start", "", "task-a", True],
        11: ["Task B", 1, dt.datetime(2026, 2, 17, 9), dt.datetime(2026, 2, 17, 10), None, dt.datetime(2026, 2, 17, 9), dt.datetime(2026, 2, 17, 10), "follow-up", "", "task-b", False],
        12: ["Phase 2", "0.5 Days", dt.datetime(2026, 2, 18, 7), dt.datetime(2026, 2, 18, 19), None, None, None, "", "", "", True],
        13: ["Task C", 3, dt.datetime(2026, 2, 18, 7), dt.datetime(2026, 2, 18, 10), None, None, None, "", "", "task-c", True],
    }
    schedule_columns = [2, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13]
    for excel_row, values in schedule_rows.items():
        for column, value in zip(schedule_columns, values, strict=True):
            ws.cell(row=excel_row, column=column, value=value)

    inputs = wb.create_sheet("Project Inputs")
    inputs.cell(row=13, column=4, value="MILL_RELINE")
    inputs.cell(row=14, column=4, value="project-123")

    metadata = wb.create_sheet("metadata")
    for column, value in enumerate(
        [
            "site-1",
            "Site Name",
            "mill-1",
            "Mill Name",
            "Vendor",
            "Liner System",
            "campaign-1",
            "Full",
            "Liner Type",
            "Supervisor",
            "Notes",
        ],
        start=1,
    ):
        metadata.cell(row=2, column=column, value=value)

    shift_definition = wb.create_sheet("shift_definition")
    shift_definition.append(["id", "project_id", "day_start_time", "night_start_time", "shift_length_hours", "timezone"])
    shift_definition.append(["shift-def-1", "project-123", dt.time(7, 0), dt.time(19, 0), 12, "America/Vancouver"])

    shift_assignments = wb.create_sheet("shift_assignments")
    shift_assignments.append(["id", "project_id", "shift_type", "crew_id", "start_date", "end_date"])
    shift_assignments.append(["assign-1", "project-123", "day", "crew-a", dt.date(2026, 2, 17), dt.date(2026, 2, 18)])
    shift_assignments.append(["assign-2", "project-123", "night", "crew-b", dt.date(2026, 2, 17), dt.date(2026, 2, 18)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def excel_bytes() -> bytes:
    return _build_excel_fixture()


def test_load_excel_project_builds_schedule_from_single_workbook_fixture(excel_bytes: bytes) -> None:
    project, metadata = ExcelProjectLoader.load_excel_project(
        file=excel_bytes,
        params=_default_excel_parameters(),
        infer_predecessors=False,
    )

    assert project.name == "Fixture Project"
    assert project.uuid == "project-123"
    assert project.project_type == ProjectType.MILL_RELINE
    assert metadata is not None
    assert metadata.site_name == "Site Name"
    assert len(project.phase_order) == 2
    assert len(project.get_task_list()) == 3
    assert project.shift_definition.project_id == "project-123"
    assert str(project.shift_definition.timezone) == "America/Vancouver"
    assert len(project.shift_assignments) == 2

    first_phase = project.phases[project.phase_order[0]]
    second_phase = project.phases[project.phase_order[1]]
    first_task = first_phase.tasks[first_phase.task_order[0]]
    second_task = first_phase.tasks[first_phase.task_order[1]]
    third_task = second_phase.tasks[second_phase.task_order[0]]

    assert first_phase.name == "Phase 1"
    assert second_phase.name == "Phase 2"
    assert first_task.name == "Task A"
    assert second_task.name == "Task B"
    assert third_task.name == "Task C"
    assert first_task.predecessor_ids == []
    assert second_task.predecessor_ids == []
    assert second_task.planned is False
    assert second_task.status == "COMPLETE"
    assert all(task.timezone_aware for task in project.get_task_list())


def test_load_excel_project_coerces_task_datetime_fields_to_python_datetimes(excel_bytes: bytes) -> None:
    project, _ = ExcelProjectLoader.load_excel_project(
        file=excel_bytes,
        params=_default_excel_parameters(),
        infer_predecessors=False,
    )

    tasks = project.get_task_list()
    assert tasks

    for task in tasks:
        assert isinstance(task.start_date, dt.datetime)
        assert isinstance(task.end_date, dt.datetime)
        if task.actual_start is not None:
            assert isinstance(task.actual_start, dt.datetime)
        if task.actual_end is not None:
            assert isinstance(task.actual_end, dt.datetime)


def test_infer_predecessors_resolves_task_and_phase_dependencies(
    excel_bytes: bytes,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fake_formula_map(*args, **kwargs) -> dict[int, str | None]:
        return {
            11: "=E10",
            12: "=E9",
        }

    monkeypatch.setattr(
        ExcelProjectLoader,
        "_extract_start_formula_map",
        staticmethod(fake_formula_map),
    )

    analysis = ExcelProjectLoader.analyze_excel_project(
        file=excel_bytes,
        params=_default_excel_parameters(),
        infer_predecessors=True,
        preview_limit=10,
    )
    project, _ = ExcelProjectLoader.load_excel_project(
        file=excel_bytes,
        params=_default_excel_parameters(),
        infer_predecessors=True,
    )

    assert analysis["provided_predecessor_count"] == 0
    assert analysis["inferred_predecessor_count"] == 1
    assert analysis["inferred_phase_predecessor_count"] == 1

    first_phase = project.phases[project.phase_order[0]]
    second_phase = project.phases[project.phase_order[1]]
    task_a = first_phase.tasks[first_phase.task_order[0]]
    task_b = first_phase.tasks[first_phase.task_order[1]]

    assert task_b.predecessor_ids == [task_a.uuid]
    assert second_phase.predecessor_ids == [first_phase.uuid]
