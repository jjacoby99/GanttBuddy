from __future__ import annotations
from dataclasses import dataclass, field
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from models.phase import Phase


@dataclass
class PhaseFormat:
    font: Font = field(default_factory=lambda: Font(name='Calibri', size=12, bold=True))
    fill: PatternFill = field(default_factory=lambda: PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")) # grey
    alignment: Alignment = field(default_factory=lambda: Alignment(horizontal="center", vertical="center"))
    row_height: int = 32.1

@dataclass
class TaskFormat:
    font: Font = field(default_factory=lambda: Font(name='Calibri', size=12, bold=False))
    fill: PatternFill = field(default_factory=lambda: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")) # white
    alignment: Alignment = field(default_factory=lambda: Alignment(horizontal="center", vertical="center"))
    row_height: int = 15.6

@dataclass
class ExcelFormat:
    # 1 indexed mapping of column names to excel column numbers
    columns: dict[str, int] = field(default_factory=lambda: {
        "number": 1, 
        "name": 2,
        "planned_duration": 3,
        "planned_start": 4,
        "planned_end": 5,
        "id": 6,
        "actual_duration": 7,
        "actual_start": 8,
        "actual_end": 9,
        "notes": 10,
        "predecessors": 11,
        "uuid": 12
    })

    first_phase_row: int = 9 # 1 indexed
    name_date_cell = "A5"

    datetime_format = "YYYY-MM-DD HH:MM"

    sheet_name: str = "Daily Schedule"

    task_format: TaskFormat = field(default_factory=TaskFormat)
    phase_format: PhaseFormat = field(default_factory=PhaseFormat)

    @property
    def date_columns(self) -> list[int]:
        return [
            self.columns["planned_start"], 
            self.columns["planned_end"],
            self.columns["actual_start"], 
            self.columns["actual_end"]
        ]