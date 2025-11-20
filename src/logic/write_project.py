import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from models.project import Project
from models.phase import Phase
from models.task import Task
from models.excel_format import ExcelFormat


class ExcelProject:
    project: Project
    excel_format: ExcelFormat
    cur_row: int
    file_path: str

    def __init__(self, project: Project, path: str, excel_format: ExcelFormat = ExcelFormat()):
        self.project = project
        self.file_path = path
        self.excel_format = excel_format
        self.cur_row = excel_format.first_phase_row # 1 indexed

    def write_project(self):
        
        wb = openpyxl.load_workbook(Path(__file__).parent.parent / "assets" / "Gantt_Excel_Template.xlsx")

        # main title cell - contains name and date range
        name_date_cell = wb[self.excel_format.sheet_name][self.excel_format.name_date_cell]
        name_date_cell.value = f"{self.project.name.split('\n')[0]}\n{self.project.start_date.date()} to {self.project.end_date.date()} (from GanttBuddy)"
        name_date_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        name_date_cell.font = openpyxl.styles.Font(name='Calibri', size=20, bold=True)

        # project name, appears above first phase row
        name_cell = wb[self.excel_format.sheet_name][f"{get_column_letter(self.excel_format.columns["name"])}{self.excel_format.first_phase_row - 1}"]
        name_cell.value = self.project.name.split('\n')[0]
        name_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        name_cell.font = openpyxl.styles.Font(name='Calibri', size=12, bold=True)

        # planned start column
        planned_start_cell = wb[self.excel_format.sheet_name][f"{get_column_letter(self.excel_format.columns["planned_start"])}{self.excel_format.first_phase_row - 1}"]
        planned_start_cell.value = self.project.start_date if self.project.start_date else None
        planned_start_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        planned_start_cell.font = openpyxl.styles.Font(name='Calibri', size=12, bold=True)

        # planned end column
        planned_end_cell = wb[self.excel_format.sheet_name][f"{get_column_letter(self.excel_format.columns["planned_end"])}{self.excel_format.first_phase_row - 1}"]
        planned_end_cell.value = self.project.end_date if self.project.end_date else None
        planned_end_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        planned_end_cell.font = openpyxl.styles.Font(name='Calibri', size=12, bold=True)


        for pid in self.project.phase_order:
            phase = self.project.phases[pid]
            self._write_phase_row(wb, phase)
            for tid in phase.task_order:
                task = phase.tasks[tid]
                self._write_task_row(wb, task)
        wb.save(self.file_path)
    
    def _write_phase_row(self, wb: openpyxl.Workbook, phase: Phase):
        if phase.uuid not in [uuid for uuid in self.project.phase_order]:
            raise ValueError(f"Provided phase {phase.name} does not exist in project {self.project.name}.")
        
        ws = wb[self.excel_format.sheet_name]
        ef = self.excel_format.phase_format

        # adjust row height to specified phase format
        ws.row_dimensions[self.cur_row].height = ef.row_height

        phase_row_dict = phase.to_excel_row()
        ws.cell(row=self.cur_row, column=self.excel_format.columns["name"], value=phase.name)
        for col_name, idx in self.excel_format.columns.items():
            cell = ws.cell(row=self.cur_row, column=idx)
            cell.font = ef.font
            cell.fill = ef.fill
            cell.alignment = ef.alignment

            # handle information not available in phase_row_dict
            to_write = phase_row_dict.get(col_name, None)
            if col_name == "number" or col_name == "id":
                to_write = self.project.get_phase_index(phase) + 1 # 1 indexed
            
            cell.value = to_write

        self.cur_row += 1
        
    
    def _write_task_row(self, wb: openpyxl.Workbook, task: Task):
        ws = wb[self.excel_format.sheet_name]
        ef = self.excel_format.task_format

        # adjust row height to specified task format
        ws.row_dimensions[self.cur_row].height = ef.row_height

        task_row_dict = task.to_excel_row()
        for col_name, idx in self.excel_format.columns.items():
            cell = ws.cell(row=self.cur_row, column=idx)
            cell.font = ef.font
            cell.fill = ef.fill
            cell.alignment = ef.alignment

            # handle information not available in task_row_dict
            to_write = task_row_dict.get(col_name, None)
            if col_name == "number" or col_name == "id":
                task_idx = self.project.get_task_idx(task)
                to_write = task_idx + 1 # 1 indexed
            
            cell.value = to_write
        
        self.cur_row += 1 



            