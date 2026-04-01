import datetime as dt
import json
import math
import os
import re
import warnings
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Optional, Union

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from logic.generate_id import new_id as new_phase_id
from models.phase import Phase
from models.project import Project
from models.project_metadata import RelineMetadata
from models.project_settings import ProjectSettings
from models.project_type import ProjectType
from models.shift_schedule import ShiftAssignment, ShiftDefinition
from models.task import Task, TaskType, new_id as new_task_id


@dataclass
class DataColumn:
    name: str
    column: int


@dataclass
class ExcelParameters:
    sheet_name: str = "Daily Schedule"

    columns: list[DataColumn] = field(default_factory=list[DataColumn])

    project_name_cell: str = "A5"
    start_row: int = 8

    def get_pd_usecols(self):
        return [col.column - 1 for col in self.columns]
    
    def get_col_names(self):
        return [col.name for col in self.columns]
    
    def get_df_index(self, col_name: str) -> int:
        for i, col in enumerate(self.columns):
            if col.name == col_name:
                return i
        raise KeyError(f"Column name '{col_name}' not found in columns.")
    
    def __getitem__(self, key: Union[int, str]) -> DataColumn:
        if isinstance(key, int):
            if key < len(self.columns):
                return self.columns[key]
            raise IndexError(f"Index {key} out of range for columns of length {len(self.columns)}.")
        elif isinstance(key, str):
            for col in self.columns:
                if col.name == key:
                    return col.column
        raise KeyError(f"Column name '{key}' not found in columns.")
    

class ProjectLoader():
    @staticmethod
    def load_json_file(file_path: str) -> dict:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File '{file_path}' not found.")
        if not os.path.splitext(file_path)[1].lower() == ".json":
            raise ValueError(f"File '{file_path}' is not a .json file.")
        
        with open(file_path, "r") as f:
            proj_dict = json.load(f)
        
        return proj_dict

    @staticmethod
    def load_project(proj_dict: dict) -> Project:
        for key in ["name", "phases", "settings"]:
            if not key in proj_dict:
                raise ValueError(f"Key '{key}' not found in project dictionary.")
            
        project = Project(name=proj_dict["name"], description=proj_dict.get("description", ""))
        
        phases = [Phase.from_dict(phase_dict) for phase_dict in proj_dict["phases"]]
        
        if len(phases) > 1:
            for i, phase in enumerate(phases[1:]):
                if phase.preceding_phase:
                    pass

        project.settings = ProjectSettings.from_dict(proj_dict["settings"])

        return project

warnings.filterwarnings(
    "ignore",
    message=".*extension is not supported and will be removed",
    module=r"openpyxl\.worksheet\._reader"
)   


@dataclass
class ExcelReadContext:
    data: bytes
    excel_file: pd.ExcelFile
    workbook_values: Workbook
    workbook_formulas: Workbook | None = None

class ExcelProjectLoader():

    _phase_pat = re.compile(r"\b(\d+(?:\.\d+)?)\s*day(s)?\b", re.IGNORECASE)
    _hours_pat = re.compile(r"\b(\d+(?:\.\d+)?)\s*hour(s)?\b", re.IGNORECASE)
    _cell_ref_pat = re.compile(
        r"^\s*=?(?:'[^']+'!)?\$?([A-Z]{1,3})\$?(\d+)\s*$",
        re.IGNORECASE,
    )

    @staticmethod
    def _is_nan(x: Any) -> bool:
        return x is None or (isinstance(x, float) and math.isnan(x)) or (isinstance(x, pd._libs.missing.NAType))
    
    @staticmethod
    def _coerce_str(x: Any) -> str:
        if ExcelProjectLoader._is_nan(x):
            return ""
        return str(x).strip()
    
    @staticmethod
    def _coerce_bool(x: Any) -> bool:
        """
        Coerce a pandas-read Excel cell value to a strict Python bool.

        Rules:
        - NaN/blank -> True (matches your default behavior)
        - bool -> itself
        - ints/floats -> 0 is False, non-zero is True
        - strings -> accepts common true/false tokens (case/whitespace-insensitive)
        True tokens:  "true", "t", "yes", "y", "1", "on"
        False tokens: "false", "f", "no", "n", "0", "off"
        - anything else -> ValueError (fail fast so bad Excel data doesn't silently pass)
        """
        if ExcelProjectLoader._is_nan(x):
            return True

        # Already a Python bool (note: bool is a subclass of int, so do this before int checks)
        if isinstance(x, bool):
            return x

        # Numeric (covers numpy numeric types too)
        if isinstance(x, (int, float)):
            # handle weird floats like 0.0, 1.0
            return bool(int(x))

        # Strings (common when Excel column has mixed types or user-entered text)
        if isinstance(x, str):
            s = x.strip().lower()
            if s == "":
                return True  # treat empty cell as default True
            if s in {"true", "t", "yes", "y", "1", "on"}:
                return True
            if s in {"false", "f", "no", "n", "0", "off"}:
                return False
            raise ValueError(f"Invalid boolean string: {x!r}")

        # Pandas sometimes gives Timestamp/other objects if the column is messed up
        raise ValueError(f"Cannot coerce to bool from type={type(x).__name__}, value={x!r}")

    @staticmethod
    def is_phase_cell(cell: Any) -> bool:
        """
        A row is a phase if PLANNED DURATION looks like a 'Days' string, not a plain number.
        """
        if ExcelProjectLoader._is_nan(cell):
            return False
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            return False  # plain numeric => task
        s = ExcelProjectLoader._coerce_str(cell)
        if not s:
            return False
        # If it mentions 'day'/'days' anywhere, treat as phase.
        return bool(ExcelProjectLoader._phase_pat.search(s))
    
    @staticmethod
    def parse_duration_hours(cell: Any) -> Optional[float]:
        """
        - If numeric -> hours directly
        - If string like '3 Days (72 Hours)' -> prefer Days*24; else use Hours if present; else None
        """
        if ExcelProjectLoader._is_nan(cell):
            return None
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            return float(cell)

        s = ExcelProjectLoader._coerce_str(cell)
        if not s:
            return None

        m_days = ExcelProjectLoader._phase_pat.search(s)
        if m_days:
            days = float(m_days.group(1))
            return days * 24.0

        m_hours = ExcelProjectLoader._hours_pat.search(s)
        if m_hours:
            return float(m_hours.group(1))

        # fallback: try to parse a bare number from string (sometimes cells are "  8  ")
        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def _infer_task_type(task_str: str) -> TaskType:
        check = task_str.lower()
        if "inch" in check:
            return TaskType.INCH
        
        if "strip" in check or "remove" in check:
            return TaskType.STRIP
        
        if "install" in check:
            return TaskType.INSTALL
        
        return TaskType.GENERIC

    def _infer_columns(wb: Workbook, params: ExcelParameters):
        """
            Reads the header row of the provided Workbook
            Infers data columns present
        """

        ws = wb[params.sheet_name]

    @staticmethod
    def _read_excel_bytes(file) -> bytes:
        if file is None:
            raise ValueError("No file provided.")

        if isinstance(file, bytes):
            return file

        if hasattr(file, "getvalue"):
            data = file.getvalue()
            if data:
                return data

        if hasattr(file, "seek"):
            file.seek(0)
        data = file.read()
        if hasattr(file, "seek"):
            file.seek(0)
        return data

    @staticmethod
    def _build_read_context(
        data: bytes,
        include_formulas: bool = False,
    ) -> ExcelReadContext:
        return ExcelReadContext(
            data=data,
            excel_file=pd.ExcelFile(BytesIO(data), engine="openpyxl"),
            workbook_values=load_workbook(BytesIO(data), read_only=True, data_only=True),
            workbook_formulas=(
                load_workbook(BytesIO(data), read_only=True, data_only=False)
                if include_formulas
                else None
            ),
        )

    @staticmethod
    def _load_schedule_dataframe(
        source: bytes | pd.ExcelFile,
        params: ExcelParameters,
    ) -> pd.DataFrame:
        df = pd.read_excel(
            source,
            sheet_name=params.sheet_name,
            header=params.start_row - 2,
            usecols=params.get_pd_usecols(),
        )

        df.columns = params.get_col_names()
        df["_excel_row"] = list(range(params.start_row, params.start_row + len(df)))
        df = df.drop(0)

        df["NOTES"] = df["NOTES"].fillna("")
        df["ACTUAL START"] = df["ACTUAL START"].replace({pd.NaT: None})
        df["ACTUAL END"] = df["ACTUAL END"].replace({pd.NaT: None})
        df["ACTIVITY"] = df["ACTIVITY"].map(ExcelProjectLoader._coerce_str)
        for col in ("PLANNED START", "PLANNED END", "ACTUAL START", "ACTUAL END"):
            df[col] = pd.to_datetime(df[col], errors="coerce")
        df = df[df["ACTIVITY"] != ""].reset_index(drop=True)
        return df

    @staticmethod
    def _parse_predecessor_cell(cell: Any) -> list[str]:
        if ExcelProjectLoader._is_nan(cell):
            return []
        return [
            pred.strip()
            for pred in ExcelProjectLoader._coerce_str(cell).split(",")
            if pred.strip()
        ]

    @staticmethod
    def _parse_direct_cell_reference(formula: Any) -> tuple[str, int] | None:
        if not isinstance(formula, str):
            return None

        match = ExcelProjectLoader._cell_ref_pat.match(formula)
        if match is None:
            return None

        col, row = match.groups()
        return col.upper(), int(row)

    @staticmethod
    def _extract_start_formula_map(
        workbook_formulas: Workbook | None,
        params: ExcelParameters,
        excel_rows: list[int],
    ) -> dict[int, str | None]:
        if workbook_formulas is None or not excel_rows:
            return {}

        formula_sheet = workbook_formulas[params.sheet_name]
        planned_start_col = params["PLANNED START"]
        formula_by_row: dict[int, str | None] = {}

        for row_cells in formula_sheet.iter_rows(
            min_row=min(excel_rows),
            max_row=max(excel_rows),
            min_col=planned_start_col,
            max_col=planned_start_col,
        ):
            cell = row_cells[0]
            formula_by_row[cell.row] = cell.value if isinstance(cell.value, str) else None

        return formula_by_row

    @staticmethod
    def _coerce_project_datetime(
        value: Any,
        timezone: dt.tzinfo | None,
    ) -> dt.datetime | None:
        if value is None or pd.isna(value):
            return None

        parsed = value.to_pydatetime() if isinstance(value, pd.Timestamp) else value
        if not isinstance(parsed, dt.datetime):
            parsed = pd.to_datetime(parsed, errors="coerce")
            if pd.isna(parsed):
                return None
            parsed = parsed.to_pydatetime()

        if timezone is None:
            return parsed.astimezone() if parsed.tzinfo is not None else parsed

        if parsed.tzinfo is None or parsed.utcoffset() is None:
            return parsed.replace(tzinfo=timezone)
        return parsed.astimezone(timezone)

    @staticmethod
    def _build_schedule_rows(
        context: ExcelReadContext,
        params: ExcelParameters,
        infer_predecessors: bool = False,
    ) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
        df = ExcelProjectLoader._load_schedule_dataframe(context.excel_file, params)

        planned_end_col = params["PLANNED END"]
        planned_end_col_letter = get_column_letter(planned_end_col)
        formula_by_row = ExcelProjectLoader._extract_start_formula_map(
            workbook_formulas=context.workbook_formulas,
            params=params,
            excel_rows=df["_excel_row"].astype(int).tolist(),
        )

        schedule_rows: list[dict[str, Any]] = []
        task_rows_by_excel_row: dict[int, dict[str, Any]] = {}

        for row_data in df.to_dict(orient="records"):
            excel_row = int(row_data["_excel_row"])
            is_phase = ExcelProjectLoader.is_phase_cell(row_data["PLANNED DURATION (HOURS)"])
            parsed_row = {
                "excel_row": excel_row,
                "is_phase": is_phase,
                "row": row_data,
                "uuid": None,
                "start_formula": None,
                "start_formula_reference": None,
                "provided_predecessors": ExcelProjectLoader._parse_predecessor_cell(row_data["PREDECESSOR"]),
                "inferred_predecessors": [],
                "resolved_predecessors": [],
                "inferred_phase_predecessors": [],
                "resolved_phase_predecessors": [],
                "predecessor_source": "provided",
            }

            formula_value = formula_by_row.get(excel_row)
            parsed_row["start_formula"] = formula_value
            parsed_row["start_formula_reference"] = ExcelProjectLoader._parse_direct_cell_reference(formula_value)

            if not is_phase:
                uuid = ExcelProjectLoader._coerce_str(row_data["UUID"])
                if uuid == "":
                    uuid = new_task_id()
                parsed_row["uuid"] = uuid
                task_rows_by_excel_row[excel_row] = parsed_row
            else:
                parsed_row["uuid"] = new_phase_id()

            schedule_rows.append(parsed_row)

        phase_rows_by_excel_row = {
            parsed_row["excel_row"]: parsed_row
            for parsed_row in schedule_rows
            if parsed_row["is_phase"]
        }

        for parsed_row in schedule_rows:
            provided_predecessors = parsed_row["provided_predecessors"]
            resolved_predecessors = list(provided_predecessors)
            predecessor_source = "provided" if provided_predecessors else "none"
            resolved_phase_predecessors: list[str] = []

            if (
                infer_predecessors
                and not parsed_row["is_phase"]
                and not provided_predecessors
            ):
                ref = parsed_row["start_formula_reference"]
                if ref is not None:
                    ref_col, ref_row = ref
                    if ref_col == planned_end_col_letter:
                        predecessor_row = task_rows_by_excel_row.get(ref_row)
                        if predecessor_row is not None and predecessor_row["uuid"] != parsed_row["uuid"]:
                            resolved_predecessors = [predecessor_row["uuid"]]
                            parsed_row["inferred_predecessors"] = list(resolved_predecessors)
                            predecessor_source = "inferred_from_start_formula"

            if infer_predecessors and parsed_row["is_phase"]:
                ref = parsed_row["start_formula_reference"]
                if ref is not None:
                    ref_col, ref_row = ref
                    if ref_col == planned_end_col_letter:
                        predecessor_phase_row = phase_rows_by_excel_row.get(ref_row)
                        if predecessor_phase_row is not None and predecessor_phase_row["uuid"] != parsed_row["uuid"]:
                            resolved_phase_predecessors = [predecessor_phase_row["uuid"]]
                            parsed_row["inferred_phase_predecessors"] = list(resolved_phase_predecessors)
                            predecessor_source = "inferred_phase_from_start_formula"

            parsed_row["resolved_predecessors"] = resolved_predecessors
            parsed_row["resolved_phase_predecessors"] = resolved_phase_predecessors
            parsed_row["predecessor_source"] = predecessor_source

        return df, schedule_rows

    @staticmethod
    def analyze_excel_project(
        file,
        params: ExcelParameters,
        infer_predecessors: bool = False,
        preview_limit: int = 100,
    ) -> dict[str, Any]:
        data = ExcelProjectLoader._read_excel_bytes(file)
        context = ExcelProjectLoader._build_read_context(
            data=data,
            include_formulas=infer_predecessors,
        )
        _, schedule_rows = ExcelProjectLoader._build_schedule_rows(
            context=context,
            params=params,
            infer_predecessors=infer_predecessors,
        )

        plan_sheet = context.workbook_values[params.sheet_name]
        project_name = plan_sheet[params.project_name_cell].value or "Untitled Project"
        project_id = ExcelProjectLoader.load_project_id(context.workbook_values)
        project_type = ExcelProjectLoader.load_project_type(context.workbook_values)
        metadata = (
            ExcelProjectLoader.load_metadata(context.workbook_values)
            if project_type == ProjectType.MILL_RELINE
            else None
        )
        shift_definition = ExcelProjectLoader.load_shift_definition(context.excel_file, project_id=project_id)
        shift_assignments = ExcelProjectLoader.load_shift_assignments(context.excel_file, project_id=project_id)

        preview_rows: list[dict[str, Any]] = []
        inferred_count = 0
        inferred_phase_count = 0
        provided_count = 0
        task_count = 0
        phase_count = 0

        for parsed_row in schedule_rows:
            row = parsed_row["row"]
            if parsed_row["is_phase"]:
                phase_count += 1
            else:
                task_count += 1

            if parsed_row["provided_predecessors"]:
                provided_count += 1
            if parsed_row["inferred_predecessors"]:
                inferred_count += 1
            if parsed_row["inferred_phase_predecessors"]:
                inferred_phase_count += 1

            preview_rows.append(
                {
                    "Excel Row": parsed_row["excel_row"],
                    "Type": "Phase" if parsed_row["is_phase"] else "Task",
                    "Activity": ExcelProjectLoader._coerce_str(row["ACTIVITY"]),
                    "Planned Duration": row["PLANNED DURATION (HOURS)"],
                    "Planned Start": row["PLANNED START"],
                    "Planned End": row["PLANNED END"],
                    "Start Formula": parsed_row["start_formula"] or "",
                    "Provided Predecessors": ", ".join(parsed_row["provided_predecessors"]),
                    "Inferred Task Predecessors": ", ".join(parsed_row["inferred_predecessors"]),
                    "Final Task Predecessors": ", ".join(parsed_row["resolved_predecessors"]),
                    "Inferred Phase Predecessors": ", ".join(parsed_row["inferred_phase_predecessors"]),
                    "Final Phase Predecessors": ", ".join(parsed_row["resolved_phase_predecessors"]),
                    "Predecessor Source": parsed_row["predecessor_source"],
                    "Planned": ExcelProjectLoader._coerce_bool(row["PLANNED"])
                    if not ExcelProjectLoader._is_nan(row["PLANNED"])
                    else True,
                }
            )

        summary = {
            "project_name": project_name,
            "project_id": project_id,
            "project_type": project_type,
            "metadata": metadata,
            "task_count": task_count,
            "phase_count": phase_count,
            "provided_predecessor_count": provided_count,
            "inferred_predecessor_count": inferred_count,
            "inferred_phase_predecessor_count": inferred_phase_count,
            "schedule_preview": pd.DataFrame(preview_rows).head(preview_limit),
            "shift_definition_preview": pd.DataFrame(
                [
                    {
                        "project_id": shift_definition.project_id,
                        "day_start_time": shift_definition.day_start_time,
                        "night_start_time": shift_definition.night_start_time,
                        "shift_length_hours": shift_definition.shift_length_hours,
                        "timezone": shift_definition.timezone,
                    }
                ]
            ),
            "shift_assignments_preview": pd.DataFrame(
                [assignment.model_dump(mode="python") for assignment in shift_assignments]
            ),
            "metadata_preview": pd.DataFrame(
                [
                    {"Field": key, "Value": value}
                    for key, value in (metadata.model_dump(mode="python").items() if metadata is not None else [])
                ]
            ),
            "column_mapping": pd.DataFrame(
                [{"Field": col.name, "Column": col.column} for col in params.columns]
            ),
        }
        return summary

    
    @staticmethod
    def load_excel_project(
        file,
        params: ExcelParameters,
        infer_predecessors: bool = False,
    ) -> tuple[Project, Optional[RelineMetadata]]:
        data = ExcelProjectLoader._read_excel_bytes(file)
        context = ExcelProjectLoader._build_read_context(
            data=data,
            include_formulas=infer_predecessors,
        )
        _, schedule_rows = ExcelProjectLoader._build_schedule_rows(
            context=context,
            params=params,
            infer_predecessors=infer_predecessors,
        )

        wb = context.workbook_values
        plan_sheet = wb[params.sheet_name]
        proj_name = plan_sheet[params.project_name_cell].value

        project = Project(name=proj_name if proj_name else "Untitled Project")

        project_id = ExcelProjectLoader.load_project_id(wb)
        if project_id is not None:
            project.uuid = project_id #existing project, preserve uuid for backend.

        project.shift_definition = ExcelProjectLoader.load_shift_definition(
            context.excel_file,
            project_id=project.uuid,
        )
        project.shift_assignments = ExcelProjectLoader.load_shift_assignments(
            context.excel_file,
            project_id=project.uuid,
        )
        if project.shift_definition is not None:
            project.timezone = project.shift_definition.timezone
        proj_type = ExcelProjectLoader.load_project_type(wb)
        project.project_type = proj_type
        metadata = None
        if proj_type == ProjectType.MILL_RELINE:
            metadata = ExcelProjectLoader.load_metadata(wb)
        current_phase = None
        unassigned_phase = None

        def mk_task(parsed_row: dict[str, Any]) -> Task:
            row = parsed_row["row"]
            name = ExcelProjectLoader._coerce_str(row["ACTIVITY"])
            return Task(
                name=name,
                start_date=ExcelProjectLoader._coerce_project_datetime(row["PLANNED START"], project.timezone),
                end_date=ExcelProjectLoader._coerce_project_datetime(row["PLANNED END"], project.timezone),
                actual_end=ExcelProjectLoader._coerce_project_datetime(row["ACTUAL END"], project.timezone),
                actual_start=ExcelProjectLoader._coerce_project_datetime(row["ACTUAL START"], project.timezone),
                note=ExcelProjectLoader._coerce_str(row["NOTES"]),
                uuid=parsed_row["uuid"],
                predecessor_ids=list(parsed_row["resolved_predecessors"]),
                planned=ExcelProjectLoader._coerce_bool(row["PLANNED"]) if not ExcelProjectLoader._is_nan(row["PLANNED"]) else True,
                task_type=ExcelProjectLoader._infer_task_type(name),
            )

        phase_ctr = 1
        task_ctr = 1
        for parsed_row in schedule_rows:
            row = parsed_row["row"]
            dur_cell = row["PLANNED DURATION (HOURS)"]
            if parsed_row["is_phase"]:
                # Commit previous phase implicitly by starting a new one
                new_phase = Phase(
                    name=ExcelProjectLoader._coerce_str(row["ACTIVITY"]),
                    uuid=parsed_row["uuid"],
                    predecessor_ids=list(parsed_row["resolved_phase_predecessors"]),
                )
                new_phase.planned = ExcelProjectLoader._coerce_bool(row["PLANNED"]) if not ExcelProjectLoader._is_nan(row["PLANNED"]) else True
                project.add_phase(new_phase)
                phase_ctr += 1
                task_ctr = 1
                current_phase = new_phase
            else:
                task = mk_task(parsed_row)
                task.infer_status()
                task_ctr += 1
                if current_phase is None:
                    # Task appears before any phase. Put it under an 'Unassigned' bucket
                    if unassigned_phase is None:
                        unassigned_phase = Phase(
                            name="Unassigned",
                            preceding_phase=None,
                        )
                        project.add_phase(unassigned_phase)
                        phase_ctr += 1
                    unassigned_phase.add_task(task)
                else:
                    current_phase.add_task(task)

        return project, metadata
    
    @staticmethod
    def load_metadata(wb, sheet_name: str="metadata") -> RelineMetadata:
        if not sheet_name in wb.sheetnames:
            raise ValueError(f"Provided sheet name {sheet_name} not found in Workbook.")
        ws = wb[sheet_name]

        site_id =          str(ws.cell(row=2,column=1).value)
        site_name =        str(ws.cell(row=2,column=2).value)
        mill_id =          str(ws.cell(row=2,column=3).value)
        mill_name =        str(ws.cell(row=2,column=4).value)
        vendor =           str(ws.cell(row=2,column=5).value)
        liner_system =     str(ws.cell(row=2,column=6).value)
        campaign_id =      str(ws.cell(row=2,column=7).value)
        scope =            str(ws.cell(row=2,column=8).value)
        liner =            str(ws.cell(row=2,column=9).value)
        supervisor =       str(ws.cell(row=2,column=10).value)
        notes =            str(ws.cell(row=2,column=11).value)
        
        return RelineMetadata(
            site_id=site_id,
            site_name=site_name,
            mill_id=mill_id,
            mill_name=mill_name,
            vendor=vendor,
            liner_system=liner_system,
            campaign_id=campaign_id,
            scope=scope,
            liner_type=liner,
            supervisor=supervisor,
            notes=notes,
        )


    @staticmethod
    def load_project_type(wb, sheet_name: str="Project Inputs") -> ProjectType:
        if not sheet_name in wb.sheetnames:
            raise ValueError(f"Provided sheet name {sheet_name} not found in Workbook.")
        
        ws = wb[sheet_name]
        
        type_str = str(ws.cell(row=13, column=4).value) # D13
        proj_type = ProjectType.GENERIC
        match type_str:
            case "MILL_RELINE":
                return ProjectType.MILL_RELINE
            case "CIVIL":
                return ProjectType.CIVIL
            case "CRUSHER_REBUILD":
                return ProjectType.CRUSHER_REBUILD
            case _:
                return ProjectType.GENERIC
        return proj_type
    
    @staticmethod
    def load_project_id(wb, sheet_name: str = "Project Inputs") -> str:
        if not sheet_name in wb.sheetnames:
            raise ValueError(f"Provided sheet name {sheet_name} not found in Workbook.")

        ws = wb[sheet_name]
        val = ws.cell(row=14, column=4).value
        if val is None:
            return None
        return str(val)

    @staticmethod
    def load_project_settings(wb, sheet_name: str = "Project Inputs") -> ProjectSettings:
        
        if not sheet_name in wb.sheetnames:
            raise ValueError(f"Provided sheet name {sheet_name} not found in Workbook.")
        
        ws = wb[sheet_name]

        work_day_start_row = 5
        work_day_col = 4

        work_days = {
            "Sunday": False,
            "Monday": False,
            "Tuesday": False,
            "Wednesday": False,
            "Thursday": False,
            "Friday": False,
            "Saturday": False
        }

        work_days_idx = {
            1: "Sunday",
            2: "Monday",
            3: "Tuesday",
            4: "Wednesday",
            5: "Thursday",
            6: "Friday",
            7: "Saturday"
        }
        # get working days
        for row in range(work_day_start_row, work_day_start_row+7):
            idx = (row - work_day_start_row + 1) % 7
            work_days[work_days_idx[idx]] = True if ws.cell(row=row,column=work_day_col).value == "yes" else False

    @staticmethod
    def coerce_time(x) -> dt.time | None:
        if pd.isna(x):
            return None

        # If it's already a python time
        if isinstance(x, dt.time) and not isinstance(x, dt.datetime):
            return x

        # If it's a datetime / pandas Timestamp -> take the time part
        if isinstance(x, (dt.datetime, pd.Timestamp)):
            return x.time()

        # Excel numeric time (fraction of day)
        if isinstance(x, (int, float)):
            # 0.5 -> 12:00:00, etc.
            total_seconds = int(round(float(x) * 24 * 60 * 60))
            total_seconds %= 24 * 60 * 60
            hh, rem = divmod(total_seconds, 3600)
            mm, ss = divmod(rem, 60)
            return dt.time(hh, mm, ss)

        # Strings like "07:30" or "7:30 AM"
        if isinstance(x, str):
            s = x.strip()
            # Try a couple common formats
            for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M:%S %p"):
                try:
                    return dt.datetime.strptime(s, fmt).time()
                except ValueError:
                    pass
            # Last resort: let pandas try
            t = pd.to_datetime(s, errors="coerce")
            if pd.isna(t):
                raise ValueError(f"Unrecognized time string: {x!r}")
            return t.time()

        raise TypeError(f"Unsupported time value type: {type(x)} ({x!r})")

    @staticmethod
    def load_shift_definition(file, project_id: str, sheet_name: str="shift_definition") -> ShiftDefinition:
        source = BytesIO(file) if isinstance(file, bytes) else file
        df = pd.read_excel(
            source,
            sheet_name=sheet_name,
            header=0,
            usecols="A:F",
            names=["id", "project_id", "day_start_time", "night_start_time", "shift_length_hours", "timezone"]
        )

        df["day_start_time"] = df["day_start_time"].map(ExcelProjectLoader.coerce_time)
        df["night_start_time"] = df["night_start_time"].map(ExcelProjectLoader.coerce_time)
        if project_id is not None and (df["project_id"] != project_id).any():
            df["project_id"] = project_id # fall back on provided project id (existing project)

        return ShiftDefinition.from_df(df, project_id)
    
    @staticmethod
    def load_shift_assignments(file, project_id: str, sheet_name: str="shift_assignments") -> list[ShiftAssignment]:
        source = BytesIO(file) if isinstance(file, bytes) else file
        df = pd.read_excel(
            source,
            sheet_name=sheet_name,
            header=0,
            usecols="A:F",
            names=["id", "project_id", "shift_type", "crew_id", "start_date", "end_date"],
        )
        df["start_date"] = pd.to_datetime(df["start_date"], errors="raise")
        df["end_date"] = pd.to_datetime(df["end_date"], errors="raise")

        return ShiftAssignment.from_df(df, project_id=project_id)

