from __future__ import annotations

from models.project import Project
from models.phase import Phase
from models.task import Task

import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import quote_sheetname
from openpyxl.worksheet.hyperlink import Hyperlink

from functools import lru_cache
from typing import Optional

class PostMortemAnalyzer:

    #@lru_cache(maxsize=128)
    @staticmethod
    def analyze_phase_delays(phase: Phase, n: int):
        """
        Takes a phase, determines the n most delayed tasks. If n == -1, the entire df is returned.
        
        Returns a Pandas DataFrame containing columns:
            "Task No."
            "Task"
            "Planned Start"
            "Planned End"
            "Planned Duration"
            "Actual Start"
            "Actual End"
            "Actual Duration"
            "Delay"
            "Notes"
        """

        data = {
            "Task No.": [],
            "Task": [],
            "Planned Start": [],
            "Planned End": [],
            "Planned Duration": [],
            "Actual Start": [],
            "Actual End": [],
            "Actual Duration": [],
            "Delay": [],
            "Notes": []
        }

        for i, tid in enumerate(phase.task_order):
            task = phase.tasks[tid]
            
            # ignore tasks with incomplete info
            if not task.completed:
                continue
            
            data["Task No."].append(i+1)
            data["Task"].append(task.name)

            pstart = task.start_date.replace(tzinfo=None)
            pend = task.end_date.replace(tzinfo=None)

            data["Planned Start"].append(pstart)
            data["Planned End"].append(pend)

            pdur = task.planned_duration.total_seconds() / 3600
            data["Planned Duration"].append(pdur)

            astart = task.actual_start.replace(tzinfo=None) if task.actual_start else None
            aend = task.actual_end.replace(tzinfo=None) if task.actual_end else None

            data["Actual Start"].append(astart)
            data["Actual End"].append(aend)


            adur = task.actual_duration.total_seconds() / 3600
            data["Actual Duration"].append(adur)
            data["Delay"].append(adur - pdur) # positive indicates a delay - negative indicates ahead of schedule

            notes = task.note
            if not notes or notes == "":
                if astart and aend:
                    notes = f"{astart} -> {aend}"
                else:
                    notes = f"{pstart} -> {pend}"
            
            data["Notes"].append(notes)
        
        df = pd.DataFrame(data)
        
        df = df.sort_values(by='Delay',ascending=False)

        if n != -1:
            return df.head(n)

        return df
    
    #@lru_cache(maxsize=128)
    @staticmethod
    def analyze_project_delays(project: Project, n: int):
        """
        Takes a project, analyzes the delays for each phase
        
        returns a list[pd.DataFrame] containing the n most delayed tasks in each Phase
        """
        delays = []
        for pid in project.phase_order:
            phase = project.phases[pid]

            delays.append(PostMortemAnalyzer.analyze_phase_delays(phase, n))
        
        return delays

    @staticmethod
    def major_delays(project: Project, n: int): 
        """
        Analyzes the project and returns a DataFrame of the top n major delays across all phases
        """
        data = {
            "Phase": [],
            "Task No.": [],
            "Task": [],
            "Delay": [],
            "Notes": []
        }

        for pid in project.phase_order:
            phase = project.phases[pid]
            phase_delay_df = PostMortemAnalyzer.analyze_phase_delays(phase, -1)

            for _, row in phase_delay_df.iterrows():
                data["Phase"].append(phase.name)
                data["Task No."].append(row["Task No."])
                data["Task"].append(row["Task"])
                data["Delay"].append(row["Delay"])
                data["Notes"].append(row["Notes"])
        
        df = pd.DataFrame(data)
        df = df.sort_values(by='Delay', ascending=False)
        if n == -1:
            return df
        return df.head(n)

    #@lru_cache(maxsize=128)
    @staticmethod
    def write_dataframe_to_sheet(df: pd.DataFrame, wb: Workbook, sheet_name: str, include_index: bool = False):

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        for row in dataframe_to_rows(df, index=include_index, header=True):
            ws.append(row)

        return wb


    @staticmethod
    def write_index_sheet(project: Project, wb: Workbook):
        ws = wb.create_sheet(
            title="Index",
            index=0
        )

        ws.append(["Sheet Name", "Description", "Link"])
        ws.append(["Major Delays", "Top delayed tasks across all phases", "Go To Sheet"])
        name_col = ws.cell(row=2, column=1, value="Major Delays")
        desc_col = ws.cell(row=2, column=2, value="Top delayed tasks across all phases")
        link_col = ws.cell(row=2, column=3, value="Go To Sheet")
        link_col.hyperlink = Hyperlink(
            ref=link_col.coordinate,
            location=f"{quote_sheetname('Major Delays')}!A1"
        )
        link_col.style = "Hyperlink"
        for i, pid in enumerate(project.phase_order):
            name = project.phases[pid].name
            num_col = ws.cell(row=3+i, column=1, value=f"Phase-{i+1}")
            name_cell = ws.cell(row=3+i, column=2, value=name)
            link_cell = ws.cell(row=3+i, column=3, value="Go To Sheet")

            sheet_name = f"Phase-{i+1}"
            if sheet_name not in wb.sheetnames:
                raise KeyError(f"Sheet {sheet_name} not found in output workbook.")
            
            link_cell.hyperlink = Hyperlink(
                ref=link_cell.coordinate,
                location=f"{quote_sheetname(sheet_name)}!A1"
            )
            link_cell.style = "Hyperlink"

        return wb

    @staticmethod
    def write_major_delays(wb: Workbook, project: Project, n: int=10):
        delays_df = PostMortemAnalyzer.major_delays(project, n)
        ws = wb.create_sheet(
            title="Major Delays",
            index=0
        )

        PostMortemAnalyzer.write_dataframe_to_sheet(
            df=delays_df,
            wb=wb,
            sheet_name="Major Delays"
        )
        return wb

    #@lru_cache(maxsize=128)
    @staticmethod
    def write_post_mortem(project: Project, n: int) -> Workbook:
        wb = Workbook() 

        wb = PostMortemAnalyzer.write_major_delays(
            wb=wb,
            project=project,
            n=n
        )

        for i, pid in enumerate(project.phase_order):
            phase = project.phases[pid]
            sheet_name = f"Phase-{i+1}"

            ws = wb.create_sheet(
                title=sheet_name,
                index=i
            ) 

            phase_delay = PostMortemAnalyzer.analyze_phase_delays(phase, n)

            wb = PostMortemAnalyzer.write_dataframe_to_sheet(
                df=phase_delay,
                wb=wb,
                sheet_name=sheet_name
            )
        try:
            wb = PostMortemAnalyzer.write_index_sheet(
                project=project,
                wb=wb
            )
        except KeyError as ex:
            print(f"Error writing index sheet for {project.name} post mortem: {str(ex)}. A sheet did not exist when constructing the index page.")

    
        return wb
    

    @staticmethod
    def cumulative_delay(project: Project) -> dict[str: float]:
        delay = {
            project.phases[pid].name: 
            (project.phases[pid].actual_end - project.phases[pid].end_date).total_seconds() / 3600
            for pid in project.phase_order
            if project.phases[pid].actual_end
        }
        
        return delay
    
    @staticmethod
    def phase_delay(phase: Phase) -> Optional[float]:
        if not phase.actual_duration:
            return None
        return (phase.actual_duration - phase.planned_duration).total_seconds() / 3600

    @staticmethod
    def get_phase_delays(project: Project) -> dict[str, float]:
        delays = {
            project.phases[pid].name:
            PostMortemAnalyzer.phase_delay(project.phases[pid])
            for pid in project.phase_order
            if project.phases[pid].actual_end and project.phases[pid].actual_start
        }
        return delays